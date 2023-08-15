import os
import sys
sys.path.append("..")
import logging

from scipy.stats import randint, uniform
from stable_baselines3 import PPO
from stable_baselines3.common.vec_env import SubprocVecEnv, VecMonitor
from stable_baselines3.common.evaluation import evaluate_policy
from modules.environments.demo_env import DemoEnvironment
from modules.models.policy import CustomActorCriticPolicy, CustomCNN
from modules.factories.callback_factory import callbacklist, CallbackType
from typing import List, Tuple
from datetime import datetime

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
import numpy as np
from stable_baselines3.common.base_class import BaseAlgorithm
from stable_baselines3.common.vec_env import VecEnv
from typing import Tuple
from apps.ml.directory_manager import DirectoryManager
import re
import platform


class ReinforcementLearningPipeline:
    @staticmethod
    def create_vectorized_environment(frequency: int, n_envs: int = os.cpu_count() or 1) -> VecMonitor:
            
        env_fns = [lambda: DemoEnvironment(rl_frequency=frequency) for _ in range(n_envs)]
        
        vectorized_environment = SubprocVecEnv(env_fns) # type: ignore
        return VecMonitor(vectorized_environment)
    
    @staticmethod
    def create_callback_list(vectorized_environment: VecMonitor, model_dir: str, log_dir: str, callbacks_to_include: List[CallbackType] = [CallbackType.EVAL, CallbackType.CHECKPOINT, CallbackType.PROGRESSBAR], n_eval_episodes: int = 10, debug: bool = False):
    
        save_freq = 100_000
        callback_list = callbacklist(
            vectorized_environment,
            log_path=log_dir,
            model_path=model_dir,
            save_freq=save_freq,
            callbacks_to_include=callbacks_to_include,
            n_eval_episodes=n_eval_episodes,
            debug=debug
        )
        return callback_list

    @staticmethod
    def create_policy_kwargs(hiddens: list) -> dict:
        policy_kwargs = dict(
            features_extractor_class=CustomCNN,
            features_extractor_kwargs=dict(features_dim=128),
            normalize_images=False,
            net_arch=dict(pi=hiddens, vf=hiddens),
        )
        return policy_kwargs
    
    @staticmethod
    def get_os_name() -> str:
        
        if platform.system() == "linux" or platform.system() == "linux2":
            return "linux"
        elif platform.system() == "posix":
            return "macos"
        elif platform.system() == 'Windows':
            return "windows"
    
        return "unknown"

    @staticmethod
    def create_ppo_model(vectorized_environment: VecMonitor, policy_kwargs: dict, learning_rate: float, debug=False) -> PPO:
        """
        Creates a PPO model with the given parameters
        CustomActorCriticPolicy is used as the policy. It receives the CustomNetwork as MLP
        the feature extractor, CustomCNN, is set in policy_kwargs and do not normalize images.
        the learning rate is set in policy_kwargs
        """
        tensorboard_log = "./logs/my_first_env/"
        device = "cuda" if ReinforcementLearningPipeline.get_os_name() == "windows" else "cpu"
        device = "mps" if ReinforcementLearningPipeline.get_os_name() == "macos" else device
        logging.info(f"Device suggested:{device}" )
        model = PPO(
            CustomActorCriticPolicy,
            vectorized_environment,
            verbose=0,
            device=device,
            tensorboard_log=tensorboard_log,
            policy_kwargs=policy_kwargs,
            learning_rate=learning_rate,
        )
        

        ReinforcementLearningPipeline.log_optimizer_learning_rate(model)
        return model

    @staticmethod
    def log_optimizer_learning_rate(model):
        optimizer_learning_rate = None

        if hasattr(model.policy, 'optimizer'):
            optimizer = model.policy.optimizer
            for param_group in optimizer.param_groups:
                if 'lr' in param_group:
                    optimizer_learning_rate = param_group['lr']
                    break
                elif 'learning_rate' in param_group:
                    optimizer_learning_rate = param_group['learning_rate']
                    break
                elif 'eta' in param_group:
                    optimizer_learning_rate = param_group['eta']
                    break

        if optimizer_learning_rate is not None:
            logging.info("Model's optimizer learning rate: %.2e", optimizer_learning_rate)
        else:
            logging.info("Model's optimizer learning rate not found.")
    
    @staticmethod
    def train_model(model: BaseAlgorithm, callback_list, n_timesteps: int = 1_000_000) -> BaseAlgorithm:
        model.learn(total_timesteps=n_timesteps, callback=callback_list)
        return model

    @staticmethod
    def gen_specific_model_folder_path(hidden_layers: List[int], rl_frequency: int, learning_rate: float, models_dir: str) -> str:
        model_folder_name = f"h{hidden_layers}-f{rl_frequency}-lr{learning_rate}"
        specific_model_dir = os.path.join(models_dir, model_folder_name)
        os.makedirs(specific_model_dir, exist_ok=True)
        return specific_model_dir
    
    #TODO change save_model input to a dictionary. it can be a kwarg
    @staticmethod
    def save_model(model: BaseAlgorithm, hidden_layers: List[int], rl_frequency: int, learning_rate: float, avg_reward: float, reward_std_dev: float, models_dir: str, debug: bool = False):
        
        specific_model_dir = ReinforcementLearningPipeline.gen_specific_model_folder_path(hidden_layers, rl_frequency, learning_rate, models_dir)
        
        model_path = os.path.join(specific_model_dir, f"m{model.__class__.__name__}-r{avg_reward}-sd{reward_std_dev}.zip")
        model.save(model_path)
        if debug:
            logging.info(f"Model saved at: {model_path}")
            
    @staticmethod
    def extract_folder_info(folder_name: str):
        pattern = r'(?P<model_name>\w+)-h\[(?P<hidden_layers>[\d, ]+)\]-f(?P<rl_frequency>\d+)-lr(?P<learning_rate>\d+\.\d+e?[-+]?\d*)-r(?P<avg_reward>\d+\.?\d*)-sd(?P<reward_std_dev>\d+\.?\d*)'
        match = re.match(pattern, folder_name)
        print(match)

        if match:
            info = match.groupdict()
            info["hidden_layers"] = list(map(int, info["hidden_layers"].split(', ')))  # Convert to a list of integers
            info["rl_frequency"] = int(info["rl_frequency"])  # Convert to an integer
            info["learning_rate"] = float(info["learning_rate"])  # Convert to a float
            info["avg_reward"] = float(info["avg_reward"])  # Convert to a float
            info["reward_std_dev"] = float(info["reward_std_dev"])  # Convert to a float
            return info
        else:
            return None
     

        
    @staticmethod
    def evaluate(model: BaseAlgorithm, env: VecEnv, n_eval_episodes: int = 100, deterministic: bool = True) -> Tuple[float, float, int]:
        """
        Evaluate the performance of a reinforcement learning model on a given environment.

        Args:
            model (BaseAlgorithm): The trained reinforcement learning agent model to be evaluated.
            env (VecEnv): The evaluation environment compatible with Stable Baselines3.
            n_eval_episodes (int, optional): The number of evaluation episodes to run. Defaults to 100.
            deterministic (bool, optional): If True, the agent will use deterministic actions during evaluation.
                                            Defaults to True.

        Returns:
            tuple: A tuple containing the average reward and standard deviation of rewards
                over the evaluation episodes.

        """
        print(f"Evaluating the model's performance over {n_eval_episodes} episodes...")

        episode_rewards, _ = evaluate_policy(model, env, render=False, n_eval_episodes=n_eval_episodes, return_episode_rewards=True, deterministic=deterministic)
        
        avg_reward = float(np.mean(episode_rewards))
        std_dev = float(np.std(episode_rewards))
        
        print(f"Average reward: {avg_reward:.2f} +/- {std_dev:.2f} over {n_eval_episodes} episodes")
        return avg_reward, std_dev, n_eval_episodes
    
    @staticmethod
    def evaluate_with_dynamic_episodes(model, env, max_episodes=200, target_std=20_000, tolerance=10_000, deterministic=True):
        """
        Evaluate the performance of a reinforcement learning model on a given environment with dynamically adjusted
        evaluation episodes.

        Args:
            model (BaseAlgorithm): The trained reinforcement learning agent model to be evaluated.
            env (gym.Env): The evaluation environment compatible with Stable Baselines3.
            max_episodes (int, optional): The maximum number of evaluation episodes to run. Defaults to 500.
            target_std (float, optional): The target standard deviation of rewards for performance convergence.
                                        Defaults to 0.1.
            tolerance (float, optional): The tolerance level for considering performance convergence. Defaults to 0.01.
            deterministic (bool, optional): If True, the agent will use deterministic actions during evaluation.
                                            Defaults to True.

        Returns:
            tuple: A tuple containing the average reward, standard deviation of rewards, and the final number of episodes
                used for evaluation.

        """
        print("Evaluating the model's performance...")

        # Começar com um número maior de episódios para uma avaliação inicial abrangente
        n_eval_episodes = 50
        num_episodes = n_eval_episodes
        all_rewards = []

        while num_episodes < max_episodes:
            print(f"Running {n_eval_episodes} evaluation episodes...")
            episode_rewards, episode_lengths = evaluate_policy(model, env, render=False, n_eval_episodes=n_eval_episodes, return_episode_rewards=True, deterministic=deterministic)
            all_rewards.extend(episode_rewards if isinstance(episode_rewards, list) else [episode_rewards])
            std_dev = np.std(all_rewards)
            num_episodes += n_eval_episodes
            print(f"conclusion: {num_episodes} episodes, std_dev: {std_dev}")
            #print("num_episodes:", num_episodes, "std_dev:", std_dev)
            
            # Verificar a convergência dentro da tolerância especificada
            if abs(std_dev - target_std) < tolerance:
                print("Performance converged within tolerance.")
                break
            else:
                print("Performance not converged yet. Running more evaluation episodes...")

        if num_episodes >= max_episodes:
            print("Maximum number of evaluation episodes reached.")
            

        avg_reward = sum(all_rewards) / len(all_rewards)
        std_dev = np.std(all_rewards)
        
        print(f"Average reward: {avg_reward:.2f} +/- {std_dev:.2f} over {len(all_rewards)} episodes")
        return avg_reward, std_dev, len(all_rewards)
    
    @staticmethod
    def load_or_create_workbook(file_path: str) -> Workbook:
        
        if os.path.isfile(file_path):
            try:
                # If the file exists, we load the workbook from the file to add the results
                workbook = load_workbook(file_path)
            except Exception as e:
                logging.error(f"Error occurred when trying to load the workbook: {e}")
                raise e
        else:
            # If the file does not exist, we create a new workbook
            workbook = Workbook()      

        if workbook.active is None:
            workbook.create_sheet()
            workbook.active = 0 
        
        return workbook
    
    @staticmethod
    def save_workbook(workbook: Workbook, file_path: str, file_name: str):
        try:
            workbook.save(file_path)
        except PermissionError:
            logging.error(f"Não foi possível salvar o arquivo '{file_name}'. Permissão negada.")
        except Exception as e:
            logging.error(f"Erro ocorrido ao salvar os resultados: {e}")    
    
    @staticmethod
    def save_results_to_excel(output_folder: str, file_name: str , results: List, headers = ['hidden_1', 'frequency', 'learning_rate', 'value']):
        file_path = os.path.join(output_folder, file_name)
        workbook: Workbook = ReinforcementLearningPipeline.load_or_create_workbook(file_path)
        sheet: Worksheet = workbook.active # type: ignore
        
        if sheet.max_row == 1 and sheet['A1'].value is None:

            sheet.append(headers)

        other_values:list = results[0:len(headers)] 
        while len(other_values) < len(headers):
            other_values.append("NaN")

        sheet.append(other_values)

        ReinforcementLearningPipeline.save_workbook(workbook, file_path, file_name)