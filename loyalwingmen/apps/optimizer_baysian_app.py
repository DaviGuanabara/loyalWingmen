import os
import sys
sys.path.append("..")
import logging
import numpy as np
import pandas as pd
from scipy.stats import randint, uniform
from stable_baselines3 import PPO
from stable_baselines3.common.vec_env import SubprocVecEnv, VecMonitor
from stable_baselines3.common.evaluation import evaluate_policy
from modules.environments.demo_env import DemoEnvironment
from modules.models.policy import CustomActorCriticPolicy, CustomCNN
from modules.factories.callback_factory import gen_eval_callback, callbacklist
from typing import List, Tuple
from datetime import datetime
import optuna
from optuna.trial import Trial
from optuna.samplers import TPESampler
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
import warnings

# Ignorar o aviso específico
#warnings.filterwarnings("ignore", message="WARN: Box bound precision lowered by casting to float32")
warnings.filterwarnings("ignore", category=UserWarning)

# Seu código aqui


# Configurando o logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')


def create_output_folder(experiment_name: str):
    # Obter a data atual
    current_time = datetime.now().strftime("%Y%m%d_%H%M%S")

    # Criar o nome da pasta com base na data atual
    #folder_name = f"output/{experiment_name}"
    folder_name = os.path.join("output", experiment_name)

    # Criar a pasta se ainda não existir
    os.makedirs(folder_name, exist_ok=True)

    # Retornar o caminho completo para a pasta
    return folder_name


def create_vectorized_environment(n_envs: int, frequency: int) -> SubprocVecEnv:
    env_fns = [DemoEnvironment for _ in range(n_envs)]
    vectorized_environment = SubprocVecEnv(env_fns)
    vectorized_environment.env_method("set_frequency", 240, frequency, indices=None)
    return vectorized_environment

def create_callback_list(n_envs: int, vectorized_environment: SubprocVecEnv, callbacks_to_include: List[str], model_dir: str, log_dir: str):
    #log_path = "./logs/"
    #model_path = "./models/"
    save_freq = 100_000
    callback_list, storage_for_callback = callbacklist(
        vectorized_environment,
        log_path=log_dir,
        model_path=model_dir,
        n_envs=n_envs,
        save_freq=save_freq,
        callbacks_to_include=callbacks_to_include
    )
    return callback_list, storage_for_callback

def create_policy_kwargs(hiddens: list, learning_rate: float) -> dict:
    policy_kwargs = dict(
        features_extractor_class=CustomCNN,
        features_extractor_kwargs=dict(features_dim=128),
        normalize_images=False,
        net_arch=dict(pi=hiddens, vf=hiddens)
    )
    return policy_kwargs

def create_model(vectorized_environment: SubprocVecEnv, policy_kwargs: dict, learning_rate: float) -> PPO:
    tensorboard_log = "./logs/my_first_env/"
    model = PPO(
        CustomActorCriticPolicy,
        vectorized_environment,
        verbose=0,
        device="auto",
        tensorboard_log=tensorboard_log,
        policy_kwargs=policy_kwargs,
        learning_rate=learning_rate,
    )
    return model

def train_model(model: PPO, callback_list, n_timesteps: int = 1_000_000) -> PPO:
    model.learn(total_timesteps=n_timesteps, callback=callback_list)
    return model

def save_model(model: PPO, hiddens: list, frequency: int, learning_rate: float, output_folder: str):
    folder_name = f"model_{hiddens}_{frequency}_{learning_rate}"
    current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
    model_dir = os.path.join(output_folder, f"models_{folder_name}_{current_time}")
    os.makedirs(model_dir, exist_ok=True)
    model_path = os.path.join(model_dir, "my_model")
    model.save(model_path)
    logging.info(f"Model saved at: {model_path}")


def cross_validation_simulation(hiddens: list, frequency: int, learning_rate: float, num_folds: int, output_folder: str, n_timesteps: int) -> float:
    number_of_logical_cores = os.cpu_count()
    n_envs = number_of_logical_cores

    vectorized_environment = create_vectorized_environment(n_envs, frequency)
    vectorized_environment = VecMonitor(vectorized_environment)
    
    #excel_folder_path = os.path.join(output_folder, 'simulation_results.xlsx')
    model_dir = os.path.join(output_folder, 'models')
    log_dir = os.path.join(output_folder, 'logs')
    callback_list, _ = create_callback_list(n_envs, vectorized_environment, callbacks_to_include=["progressbar"], model_dir=model_dir, log_dir=log_dir)
    policy_kwargs = create_policy_kwargs(hiddens, learning_rate)
    model = create_model(vectorized_environment, policy_kwargs, learning_rate)

    logging.info(model.policy)
    model = train_model(model, callback_list, n_timesteps)
    save_model(model, hiddens, frequency, learning_rate, output_folder)
    avg_score, std_dev = evaluate_policy(model, vectorized_environment, n_eval_episodes=num_folds, deterministic=True)
    logging.info(f"Avg score: {avg_score}")
    return avg_score

def generate_random_parameters() -> Tuple[list, int, float]:
    hidden_dist = randint(10, 1000)
    num_hiddens = randint(3, 8).rvs()
    hiddens = [hidden_dist.rvs() for _ in range(num_hiddens)]
    frequency = (randint(1, 8).rvs() * 15)
    learning_rate = uniform(0.00000000001, 0.1).rvs() 
    return hiddens, frequency, learning_rate





def save_results_to_excel(results, output_folder, file_name):
    # Construir o caminho completo do arquivo
    file_path = os.path.join(output_folder, file_name)

    # Verificar se o diretório de saída existe
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    # Verificar se o arquivo existe
    if not os.path.isfile(file_path):
        # Se o arquivo não existe, criar uma nova planilha
        workbook = Workbook()
        sheet: Worksheet = workbook.active
        
        # Definir os cabeçalhos das colunas
        sheet.append(['hiddens', 'frequency', 'learning_rate', 'score'])    
    else:
        # Carregar a planilha existente
        workbook = load_workbook(file_path)
        # Obter a planilha ativa
        sheet: Worksheet = workbook.active
        
    
        
    for result in results:
        hiddens_str = ', '.join(str(x) for x in result[0])
        sheet.append([hiddens_str, result[1], result[2], result[3]])

    try:
        # Salvar as alterações no arquivo
        workbook.save(file_path)
    except PermissionError:
        print(f"Não foi possível salvar o arquivo '{file_name}'. Permissão negada.")
    except Exception as e:
        print(f"Erro ocorrido ao salvar os resultados: {e}")    

    # Adicionar os novos resultados
    #for result in results:
    #    sheet.append(result)






def objective(trial: Trial, output_folder: str, n_timesteps: int) -> float:
    
    num_hiddens = trial.suggest_int('num_hiddens', 3, 8)
    hiddens = [trial.suggest_int(f'hiddens_{i}', 100, 1000) for i in range(num_hiddens)]
    frequency = trial.suggest_int('frequency', 1, 8) * 15
    exponent = trial.suggest_uniform('exponent', -10, -1)
    learning_rate = 10 ** exponent

    
    print("Parameters:")
    print("Hiddens: ", hiddens)
    print("Frequency: ", frequency)
    print("Learning Rate: ", learning_rate)

    avg_score = cross_validation_simulation(hiddens, frequency, learning_rate, num_folds=5, output_folder=output_folder, n_timesteps=n_timesteps)

    # Salvar os resultados na planilha
    result = (hiddens, frequency, learning_rate, avg_score)
    save_results_to_excel([result], output_folder, 'simulation_results.xlsx')

    return avg_score

def print_best_parameters(results: List[Tuple[List[int], int, float, float]]):
    results = sorted(results, key=lambda x: x[-1], reverse=True)
    print("Best score: %.4f" % results[0][-1])
    print("Best parameters:")
    print("Hiddens: %s" % ', '.join(map(str, results[0][0])))
    print("Frequency: %.4f" % results[0][1])
    print("Learning rate: %.10f" % results[0][2])

def main():
    
    
    n_timesteps = 1_000_000
    experiment_name = "Optimizer_baysian_app"
    # Criar a pasta de saída
    output_folder = create_output_folder(experiment_name)

    study = optuna.create_study(direction='maximize', sampler=TPESampler())
    study.optimize(lambda trial: objective(trial, output_folder, n_timesteps), n_trials=100)

    results = []
    for trial in study.trials:
        hiddens = [trial.params[f'hiddens_{i}'] for i in range(trial.params['num_hiddens'])]
        results.append((hiddens, trial.params['frequency'], trial.params['learning_rate'], trial.value))

    #save_results_to_excel(results)
    print_best_parameters(results)

if __name__ == "__main__":
    main()
